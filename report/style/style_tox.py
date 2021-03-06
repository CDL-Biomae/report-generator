from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl import load_workbook
from openpyxl.utils.cell import get_column_letter
from termcolor import colored
from tools import QueryScript

import env

def add_style_tox(tox_dataframe, PATH):

    wb = load_workbook(PATH)
    ws = wb['Tox']

    nb_rows, nb_columns = tox_dataframe.shape

    thin = Side(border_style='thin', color='FF000000')
    medium = Side(border_style='medium', color='FF000000')
    double = Side(border_style='double', color='FF000000')
    no_border = Side(border_style=None)

    font_Arial9 = Font(size=9, name='Arial')
    font_title = Font(name='Arial', size=10, bold=True, color='FF000000')
    font_small_title = Font(name='Arial', size=8, bold=False, color='FF000000')
    alignment_center = Alignment(horizontal='center', vertical='center')

    ## COLUMN WIDTH ##
    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 45
    ws.column_dimensions['E'].width = 14

    ws.row_dimensions[3].height = 45

    # Nettoyage Fin du tableau
    cell = ws['S4']
    cell.border = no_border

    # Nettoyage entre 2 tableaux
    cell = ws['F4']
    cell.border = Border(top=no_border, bottom=no_border)
    ws.column_dimensions['F'].width = 3

    medium_border = Border(left=medium, right=medium, top=medium, bottom=medium)
    normal_cells_border = Border(left=thin,   right=thin, top=thin, bottom=thin)
    header_stations_font = Font(size=10, bold=True, name='Arial', color='FFFFFF')
    header_stations_fill = PatternFill(fill_type='solid', start_color='808080', end_color='808080')
    center_alignment = Alignment(horizontal='center', vertical='center')

    ## REFORMATAGE STATIONS DATA
    columns_stations = ['B', 'C', 'D', 'E']
    for column in columns_stations:
        for row in range(5, 5+nb_rows):
            cell = ws[column + str(row)]
            if column == 'B':
                if row == 5+nb_rows-1:
                    cell.border = Border(left=medium, top=thin, right=thin, bottom=medium)
                else:
                    cell.border = Border(left=medium, top=thin, right=thin, bottom=thin)
            elif column == 'E':
                if row == 5+nb_rows-1:
                    cell.border = Border(left=thin, top=thin, right=medium, bottom=medium)
                else:
                    cell.border = Border(left=thin, top=thin, right=medium, bottom=thin)
            elif row == 5+nb_rows-1:
                cell.border = Border(left=thin, top=thin, right=thin, bottom=medium)
            else:
                cell.border = normal_cells_border

            cell.font = font_Arial9
            cell.alignment = center_alignment


    ## REFORMATAGE VALUE HEADER ##
    titres_columns = ['G', 'H', 'I', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R']
    dict_titre_columns = {'G': 'Survie Male - 7 jours',
                          'H': 'Alimentation',
                          'I': 'Neurotoxicité AChE',
                          'K': 'Survie Femelle',
                          'L': 'Nombre jours exposition in situ',
                          'M': 'n',
                          'N': 'Fécondité',
                          'O': 'n',
                          'P': 'Cycle de mue',
                          'Q': 'n',
                          'R': 'Perturbation endocrinienne'}
    subtitle = ['L', 'N', 'P', 'R', 'G', 'H', 'I', 'K']
    n_merge = ['L3:L4', 'M3:M4', 'O3:O4', 'Q3:Q4']

    ws['N4'].value = "indice"
    ws['P4'].value = "valeur observée (valeur attendue)"
    ws['R4'].value = "surface ovocytaire moyenne (µm²)"

    # Nettoyage entre 2 tableaux
    cell = ws['J4']
    cell.border = Border(top=no_border, bottom=no_border)
    ws.column_dimensions['J'].width = 3

    # Merge et style des colonnes
    for merge_column in n_merge:
        ws.merge_cells(merge_column)
        top_left_cell = ws[merge_column[0:2]]
        top_left_cell.value = dict_titre_columns[merge_column[0]]
        top_left_cell.font = font_title
        top_left_cell.border = medium_border

    # Style des titres
    for column in titres_columns:
        if column in ['L', 'P', 'R']:
            ws.column_dimensions[column].width = 35
        elif column in ['M', 'O', 'Q']:
            ws.column_dimensions[column].width = 5
        else:
            ws.column_dimensions[column].width = 20

        cell = ws[column + "3"]
        cell.value = dict_titre_columns[column]
        cell.font = font_title
        cell.alignment = center_alignment
        cell.border = medium_border

    # Style des sous-titres
    for column in subtitle:
        cell = ws[column + "4"]
        cell.font = font_small_title
        cell.border = medium_border

    # Nettoyage des borders des données cachées

    ws["T4"].border = Border(top=no_border,left=no_border, right=no_border, bottom=no_border)
    ws["u4"].border = Border(top=no_border,left=no_border, right=no_border, bottom=no_border)

    ## STYLE VALUE DATA
    columns = [get_column_letter(col_idx) for col_idx in range(6, nb_columns+2)]
    little_border_columns = ['G', 'H', 'I', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R']
    na = ['L', 'M', 'N', 'O', 'P', 'Q', 'R']

    body_fill_ok = PatternFill(fill_type='solid', start_color='027ee3', end_color='027ee3') #bleu
    body_fill_not_ok_1 = PatternFill(fill_type='solid', start_color='69a64b', end_color='69a64b') #green
    body_fill_not_ok_2 = PatternFill(fill_type='solid', start_color='d1c452', end_color='d1c452') #yellow
    body_fill_not_ok_3 = PatternFill(fill_type='solid', start_color='cc7931', end_color='cc7931') #orange
    body_fill_not_ok_4 = PatternFill(fill_type='solid', start_color='ab2222', end_color='ab2222') #red
    body_fill_NA = PatternFill(fill_type='solid', start_color='abadb0', end_color='abadb0')  # grey



    # la requete qui retourne les resultats des constantes r2 selon le paramétre
    threshold_data = QueryScript(f" SELECT parameter, threshold   FROM {env.DATABASE_TREATED}.r2_threshold WHERE threshold IS NOT NULL and version=  {env.CHOSEN_VERSION()}").execute()
    threshold_list = []
    for column in columns:
        
        threshold = None        
        if column == 'H':
            threshold = []
            for element in threshold_data:
                if element[0] == 'alimentation':
                    threshold.append(element[1])
            threshold_list.append(threshold)
        if column == 'I':
            threshold = []
            for element in threshold_data:
                if element[0] == 'neurotoxicité AChE':
                    threshold.append(element[1])
            threshold_list.append(threshold)
        if column == 'N':
            threshold = []
            for element in threshold_data:
                if element[0] == 'reproduction':
                    threshold.append(element[1])
            threshold_list.append(threshold)
        if column == 'P':
            threshold = []
            for element in threshold_data:
                if element[0] == 'mue':
                    threshold.append(element[1])
        if column == 'R':
            threshold = []
            for element in threshold_data:
                if element[0] == 'perturbation endocrinienne':
                    threshold.append(element[1])

        if threshold:
            value=None
            for row in range(5, nb_rows+5):
                cell = ws[column+str(row)]
                
                if isinstance(cell.value,float) or isinstance(cell.value,int) or (isinstance(cell.value,str) and cell.value[:2] != "NA"):
                    value = -float(cell.value)  if cell.value else None  
                    
                if cell.value != None :
                    already_done = False
                    if column == 'H' :                    
                        survival_result = float(ws['G'+str(row)].value if ws['G'+str(row)].value and ws['G'+str(row)].value!='NA' else 0)
                        if survival_result!=None and survival_result < 70 :
                            if survival_result < 50:
                                already_done = True
                            else :
                                cell.fill = body_fill_NA
                                already_done = True
                        
                    if  not already_done:
                        if value and value >= threshold[0] :
                            if len(threshold) >= 1 and value >= threshold[1]:
                                if len(threshold) >= 2 and value >= threshold[2]:
                                    if len(threshold) >= 3 and value >= threshold[3]:
                                        cell.fill = body_fill_not_ok_4
                                    else:
                                        cell.fill = body_fill_not_ok_3
                                else:
                                    cell.fill = body_fill_not_ok_2
                            else:
                                cell.fill = body_fill_not_ok_1
                        else:
                            cell.fill = body_fill_ok

        if column == 'G' or column == 'H' or column == 'I' or column == 'K':
            cell = ws[column + "4"]
            cell.value = '%'

    # pour faire les bordures des colones et ligne selon les case et les centrer
    for column in little_border_columns:
        for row in range(5, 5+nb_rows):
            ws[column + str(row)].border = normal_cells_border
            ws[column + str(row)].alignment = alignment_center

    # colorer les cycle de mue et la Perturbation endocrinienne
    for row in range(5, nb_rows+5): 

        mue= ws["T" +str(row)].value
        surface_retard = ws["U" + str(row)].value
        if ws["O" + str(row)].value :
            try :
                value = int(ws["O" + str(row)].value)
                if int(ws["O" + str(row)].value) >= 10 :
                    if ws["R" + str(row)].value == "NA" or surface_retard == "Conforme BC1":
                        ws["R" + str(row)].fill = body_fill_ok
                    else:
                        ws["R" + str(row)].fill = body_fill_not_ok_4

                    if mue == "NA":
                            ws["P" + str(row)].fill = body_fill_NA
                    if mue == "Conforme":
                            ws["P" + str(row)].fill = body_fill_ok
                    if mue == "Retard modéré":
                            ws["P" + str(row)].fill = body_fill_NA
                    if mue == "Retard fort":
                            ws["P" + str(row)].fill = body_fill_not_ok_4
                else :
                    ws["R" + str(row)].fill = body_fill_NA
                    ws["P" + str(row)].fill = body_fill_NA
                    ws["N" + str(row)].fill = body_fill_NA
            except ValueError:
                ws["R" + str(row)].fill = body_fill_NA
                ws["P" + str(row)].fill = body_fill_NA
                ws["N" + str(row)].fill = body_fill_NA
        else :
            ws["R" + str(row)].fill = body_fill_NA
            ws["P" + str(row)].fill = body_fill_NA
            ws["N" + str(row)].fill = body_fill_NA

                
        ws["S"+ str(row)].value = ""  
        ws["T" +str(row)].value = ""
        ws["U" +str(row)].value = ""



    #si la survie femelle est egale à 0 ou bien il y'a pas donner la coloration de la ligne sera grise
    for row in range(5, nb_rows+5):
        value = ws["K" + str(row)].value

        testNone = value is None
        testInt = isinstance(value, int) and float(value) == 0
        testStr = isinstance(value, str) and value[:4] == 'None'
        if testNone or testInt or testStr:
            for column in na:
                ws[column + str(row)].value = "NA"
                if column == "N" or column == "P" or column == "R":
                    ws[column + str(row)].fill = body_fill_NA
    ws.freeze_panes = ws["F8"]

    wb.save(PATH)
    wb.close()

    wb = load_workbook(PATH)
    ws = wb['Tox']
    ## Add threshold table
    ws.insert_rows(1)
    ws.insert_rows(1)
    ws.insert_rows(1)

    ws["F1"].value = "Effet faible"
    ws["F2"].value = "Effet modéré"
    ws["F3"].value = "Effet élevé"
    ws["F4"].value = "Effet très élevé"
    ws.column_dimensions['F'].width=20
    header_columns = [get_column_letter(col_idx) for col_idx in list(range(2, nb_columns + 2))]
    for letter in header_columns[5:]:
        index = None
        if letter=='H':
            index=0
        if letter=='I':
            index=1
        if letter=='N':
            index=2
        if index!=None :
            [threshold, graduate_25, graduate_50, graduate_75] = threshold_list[index]
            ws[letter+"1"].value = -threshold
            ws[letter+"1"].border = medium_border
            ws[letter+"2"].value = -graduate_25
            ws[letter+"2"].border = medium_border
            ws[letter+"3"].value = -graduate_50
            ws[letter+"3"].border = medium_border
            ws[letter+"4"].value = -graduate_75
            ws[letter+"4"].border = medium_border

    ## REFORMATAGE STATIONS HEADER ##
    merging_cells = ['B5:B7', 'C5:C7', 'D5:D7', 'E5:E7']
    border_cells = ['B5', 'B7', 'C5', 'C7', 'D5', 'D7', 'E5', 'E7']

    merging_names = ['Campagne', 'Numéro', 'Station de mesure', 'Code Agence']

    # merger les cellules selon la taille de l'entete d'excel
    for i in range(len(merging_cells)):
        cells = merging_cells[i]
        top_left_cell = ws[cells[0:2]]
        name = merging_names[i]

        ws.merge_cells(cells)
        top_left_cell.value = name
        top_left_cell.font = header_stations_font
        top_left_cell.fill = header_stations_fill
        top_left_cell.alignment = center_alignment

    for cell in border_cells:
        ws[cell].border = medium_border

   
    wb.save(PATH)
    wb.close()